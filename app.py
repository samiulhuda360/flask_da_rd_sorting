from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
from openpyxl import Workbook
from urllib.parse import urlparse
from openpyxl.utils.cell import get_column_letter

import csv
import io

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Set a secret key for Flash messages
# Define the rating ranges
rating_ranges = [
    (0, 10),
    (11, 20),
    (21, 30),
    (31, 40),
    (41, 50),
    (51, 60),
    (61, 70),
    (71, 80),
    (81, 90),
    (91, 100)
]

# Define the referring domains ranges
rd_ranges = [
    (1, 100),
    (101, 200),
    (201, 300),
    (301, 400),
    (401, 500),
    (501, 600),
    (601, 700),
    (701, 800),
    (801, 900),
    (901, 1000),
    (1001, float('inf'))  # 1000+
]

@app.route('/domain-sorting', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        uploaded_files = request.files.getlist('files[]')

        if not uploaded_files:
            flash('No file chosen. Please select at least one file.', 'error')
            return redirect(url_for('index'))

        if len(uploaded_files) > 5:
            flash('Too many files uploaded. Maximum allowed is 5.', 'error')
            return redirect(url_for('index'))

        rating_counts = {}

        for file in uploaded_files:
            if file.filename.endswith('.csv'):
                file_content = file.stream.read().decode('utf-8')
                csv_reader = csv.reader(file_content.splitlines())

                # Read the header row
                header_row = next(csv_reader)

                # Find the index of the "Domain rating", "Target URL", and "Referring domains" columns
                domain_rating_column_index = None
                target_url_column_index = None
                referring_domains_column_index = None
                for index, column_name in enumerate(header_row):
                    if "Domain rating" in column_name:
                        domain_rating_column_index = index
                    elif "Target URL" in column_name:
                        target_url_column_index = index
                    elif "Referring domains" in column_name:
                        referring_domains_column_index = index

                if domain_rating_column_index is None or target_url_column_index is None or referring_domains_column_index is None:
                    flash(f"Required columns not found in {file.filename}.", 'error')
                    return redirect(url_for('index'))

                # Read the first data row to get the competitor name
                try:
                    first_row = next(csv_reader)
                    target_url = first_row[target_url_column_index]
                    competitor_name = urlparse(target_url).netloc
                except (StopIteration, IndexError):
                    flash(f"No data found in {file.filename}.", 'error')
                    return redirect(url_for('index'))

                # Initialize the count for the competitor if it doesn't exist
                if competitor_name not in rating_counts:
                    rating_counts[competitor_name] = {
                        'domain_rating': {range_: 0 for range_ in rating_ranges},
                        'referring_domains': {range_: 0 for range_ in rd_ranges}
                    }

                # Iterate over each row in the CSV (including the first row)
                for row_index, row in enumerate(csv_reader, start=1):
                    try:
                        domain_rating = float(row[domain_rating_column_index])
                        referring_domains = int(row[referring_domains_column_index])

                        # Increment the count for the corresponding rating range
                        for range_ in rating_ranges:
                            if range_[0] <= domain_rating <= range_[1]:
                                rating_counts[competitor_name]['domain_rating'][range_] += 1
                                break

                        # Increment the count for the corresponding referring domains range
                        for range_ in rd_ranges:
                            if range_[0] <= referring_domains <= range_[1]:
                                rating_counts[competitor_name]['referring_domains'][range_] += 1
                                break
                    except (IndexError, ValueError) as e:
                        flash(f"Error processing row {row_index + 1} in {file.filename}: {e}", 'error')
                        return redirect(url_for('index'))

        if not rating_counts:
            flash('No data to generate report.', 'error')
            return redirect(url_for('index'))

        # Create a new workbook and select the active sheet
        workbook = Workbook()
        sheet = workbook.active

        # Write the header row for domain rating counts
        sheet.cell(row=1, column=1, value="Competitor")
        for col, range_ in enumerate(rating_ranges, start=2):
            sheet.cell(row=1, column=col, value=f"DR {range_[0]}-{range_[1]}")

        # Write the data rows for domain rating counts
        row_idx = 2
        for competitor_name in rating_counts.keys():
            sheet.cell(row=row_idx, column=1, value=competitor_name)
            for col, range_ in enumerate(rating_ranges, start=2):
                count = rating_counts[competitor_name]['domain_rating'][range_]
                sheet.cell(row=row_idx, column=col, value=count)
            row_idx += 1

        # Calculate and write the average for domain rating counts
        sheet.cell(row=row_idx, column=1, value="Average")
        for col, range_ in enumerate(rating_ranges, start=2):
            total_count = sum(rating_counts[competitor_name]['domain_rating'][range_] for competitor_name in rating_counts.keys())
            avg_count = total_count / len(rating_counts)
            sheet.cell(row=row_idx, column=col, value=avg_count)

        # Write the header row for referring domains counts
        start_row = row_idx + 2
        sheet.cell(row=start_row, column=1, value="Competitor")
        for col, range_ in enumerate(rd_ranges, start=2):
            sheet.cell(row=start_row, column=col, value=f"RD {range_[0]}-{range_[1]}")

        # Write the data rows for referring domains counts
        row_idx = start_row + 1
        for competitor_name in rating_counts.keys():
            sheet.cell(row=row_idx, column=1, value=competitor_name)
            for col, range_ in enumerate(rd_ranges, start=2):
                count = rating_counts[competitor_name]['referring_domains'][range_]
                sheet.cell(row=row_idx, column=col, value=count)
            row_idx += 1

        # Calculate and write the average for referring domains counts
        sheet.cell(row=row_idx, column=1, value="Average")
        for col, range_ in enumerate(rd_ranges, start=2):
            total_count = sum(rating_counts[competitor_name]['referring_domains'][range_] for competitor_name in rating_counts.keys())
            avg_count = total_count / len(rating_counts)
            sheet.cell(row=row_idx, column=col, value=avg_count)

        # Adjust the column width
        for col_num, col_letter in enumerate(get_column_letter(2), start=2):
            sheet.column_dimensions[col_letter].width = 15

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name='domain_rating_report.xlsx')

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)